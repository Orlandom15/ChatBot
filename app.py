from flask import Flask, render_template, request, jsonify, Response
import uuid
import os
from datetime import datetime
from config.database import NeonDatabase
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')


app = Flask(__name__)
db = NeonDatabase()

# Configuración específica para Render
if os.environ.get('RENDER'):
    # Render usa puerto 10000 internamente
    app.config['SERVER_NAME'] = os.environ.get('RENDER_EXTERNAL_HOSTNAME', 'localhost')
    
    # Configuración de seguridad para producción
    app.config['SESSION_COOKIE_SECURE'] = True
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    
    # Deshabilitar debug en producción
    app.config['DEBUG'] = False
    app.config['TESTING'] = False

# ============================================
# FUNCIONES PARA REEMPLAZAR pandas
# ============================================

def generar_excel_sin_pandas(data, nombre_hoja="Datos"):
    """
    Reemplaza pd.DataFrame().to_excel() completamente
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    
    if not data:
        return wb
    
    # Obtener encabezados
    headers = list(data[0].keys())
    
    # Escribir encabezados con estilo
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = item.get(header)
            # Convertir tipos especiales
            if isinstance(value, bool):
                value = '✅ Sí' if value else '❌ No'
            elif isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def descargar_excel_como_bytes(data, filename="reporte.xlsx"):
    """Descarga Excel como bytes (reemplaza pandas)"""
    wb = generar_excel_sin_pandas(data)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def generar_pdf_sin_pandas(data, titulo="Reporte"):
    """Generar PDF sin usar pandas"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph(titulo, styles['Title'])
    elements.append(title)
    
    # Crear tabla directamente desde los datos
    if data:
        # Encabezados
        headers = list(data[0].keys())
        table_data = [headers]
        
        # Datos
        for item in data:
            row = []
            for header in headers:
                value = item.get(header, '')
                # Formatear valores especiales
                if isinstance(value, bool):
                    value = '✅ Sí' if value else '❌ No'
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                row.append(str(value))
            table_data.append(row)
        
        # Crear tabla
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# ============================================
# ENDPOINTS
# ============================================

# Health check mejorado para Render
@app.route('/health')
def health_check():
    """Health check endpoint para Render"""
    try:
        # Verificar base de datos
        db.test_connection()
        
        health_status = {
            'status': 'healthy',
            'service': 'ChatBot Universitario',
            'timestamp': datetime.now().isoformat(),
            'version': '1.0.0',
            'environment': os.environ.get('FLASK_ENV', 'development'),
            'database': 'connected',
            'endpoints': {
                'chat': '/chat',
                'metrics': '/metrics',
                'api_docs': '/api'
            }
        }
        
        return jsonify(health_status)
        
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 503

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/chat', methods=['POST'])
def chat():
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        session_id = data.get('session_id')
        
        if not user_message:
            return jsonify({'success': False, 'error': 'Mensaje vacío'})
        
        print(f"💬 Mensaje: '{user_message}' - Sesión: {session_id}")
        
        # Obtener respuesta del bot
        bot_response, intent, confidence = db.get_bot_response(user_message)
        
        # Preparar respuesta base
        response_data = {
            'success': True,
            'bot_response': bot_response,
            'intent': intent,
            'confidence': confidence
        }
        
        # 🔥 AGREGAR DATOS ESTRUCTURADOS SEGÚN EL TIPO DE CONSULTA
        user_lower = user_message.lower()
        
        # Consulta de estadísticas
        if any(palabra in user_lower for palabra in ['estadística', 'estadisticas', 'estadistica', 'total', 'cuántos', 'cuantos']):
            print("📊 Obteniendo estadísticas universitarias...")
            estadisticas = db.get_estadisticas_estudiantes()
            response_data['estadisticas'] = estadisticas
            print(f"✅ Estadísticas obtenidas: {estadisticas}")
        
        # Consulta de estudiantes pendientes
        elif any(palabra in user_lower for palabra in ['pendiente', 'pendientes', 'debe', 'inscripción', 'inscripcion', 'pago']):
            print("📋 Obteniendo estudiantes pendientes...")
            estudiantes = db.get_estudiantes_pendientes_inscripcion()
            response_data['estudiantes'] = estudiantes
            print(f"✅ Estudiantes pendientes obtenidos: {len(estudiantes)}")
        
        # Consulta de carreras
        elif any(palabra in user_lower for palabra in ['carrera', 'carreras', 'ingeniería', 'sistemas', 'industrial']):
            print("🎓 Obteniendo carreras...")
            carreras = db.get_carreras()
            response_data['carreras'] = carreras
            print(f"✅ Carreras obtenidas: {len(carreras)}")
        
        # Consulta de reportes
        elif any(palabra in user_lower for palabra in ['reporte', 'archivo', 'descargar', 'generar']):
            print("📄 Generando reporte COMPLETO de estudiantes...")
            
            # ✅ Usar el NUEVO método que obtiene TODOS los estudiantes
            reporte = db.generar_reporte_completo_estudiantes()
            response_data['reporte'] = reporte
            print(f"✅ Reporte COMPLETO generado: {reporte.get('total_estudiantes', 0)} estudiantes")
        
        # Guardar conversación
        if session_id:
            db.save_conversation(
                session_id=session_id,
                user_message=user_message,
                bot_response=bot_response,
                intent=intent,
                confidence=confidence
            )
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Error en /chat: {e}")
        import traceback
        print(f"🔍 Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': 'Error interno',
            'bot_response': 'Lo siento, hubo un error. Por favor intenta nuevamente.'
        })

@app.route('/history')
def get_history():
    try:
        session_id = request.args.get('session_id')
        
        if not session_id:
            return jsonify({'success': True, 'history': []})
        
        chat_history = db.get_chat_history(session_id, limit=20)
        
        formatted_history = []
        for msg in chat_history:
            formatted_history.append({
                'type': msg['message_type'],
                'message': msg['user_message'] if msg['message_type'] == 'user' else msg['bot_response'],
                'timestamp': msg['created_at'].isoformat() if msg['created_at'] else datetime.now().isoformat()
            })
        
        return jsonify({
            'success': True,
            'history': formatted_history
        })
        
    except Exception as e:
        print(f"❌ Error en /history: {e}")
        return jsonify({'success': False, 'error': 'Error obteniendo historial'})

# 🔥 NUEVAS RUTAS PARA DATOS UNIVERSITARIOS
@app.route('/api/universidad/estadisticas')
def get_estadisticas_universidad():
    """Endpoint directo para obtener estadísticas"""
    try:
        estadisticas = db.get_estadisticas_estudiantes()
        return jsonify({
            'success': True,
            'estadisticas': estadisticas
        })
    except Exception as e:
        print(f"❌ Error en /api/universidad/estadisticas: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/universidad/estudiantes/pendientes')
def get_estudiantes_pendientes():
    """Endpoint directo para obtener estudiantes pendientes"""
    try:
        estudiantes = db.get_estudiantes_pendientes_inscripcion()
        return jsonify({
            'success': True,
            'estudiantes': estudiantes,
            'total': len(estudiantes)
        })
    except Exception as e:
        print(f"❌ Error en /api/universidad/estudiantes/pendientes: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/universidad/carreras')
def get_carreras_universidad():
    """Endpoint directo para obtener carreras"""
    try:
        carreras = db.get_carreras()
        return jsonify({
            'success': True,
            'carreras': carreras
        })
    except Exception as e:
        print(f"❌ Error en /api/universidad/carreras: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/universidad/estudiantes/todos')
def get_todos_estudiantes():
    """Endpoint directo para obtener TODOS los estudiantes"""
    try:
        limit = request.args.get('limit', 200, type=int)
        estudiantes = db.get_todos_estudiantes(limit)
        
        print(f"📊 Obtenidos {len(estudiantes)} estudiantes (todos)")
        
        return jsonify({
            'success': True,
            'estudiantes': estudiantes,
            'total': len(estudiantes),
            'limit': limit
        })
    except Exception as e:
        print(f"❌ Error en /api/universidad/estudiantes/todos: {e}")
        return jsonify({'success': False, 'error': str(e)})

# 📊 NUEVAS RUTAS PARA DESCARGAS (SIN pandas)
@app.route('/descargar/excel')
def descargar_excel():
    """Descargar reporte de estudiantes en Excel SIN pandas"""
    try:
        # Obtener todos los estudiantes
        estudiantes = db.get_todos_estudiantes(limit=1000)
        
        if not estudiantes:
            return jsonify({'success': False, 'error': 'No hay estudiantes para exportar'})
        
        # Preparar datos
        data = []
        for est in estudiantes:
            data.append({
                'Matrícula': est['matricula'],
                'Nombre': f"{est['nombre']} {est['apellido']}",
                'Carrera': est['carrera'],
                'Semestre': est['semestre'],
                'Fecha Inscripción': est['fecha_inscripcion'].strftime('%Y-%m-%d') if est['fecha_inscripcion'] else 'N/A',
                'Estado Pago': '✅ PAGADO' if est.get('inscripcion_pagada') else '❌ PENDIENTE',
                'Email': est.get('email', 'No especificado'),
                'Teléfono': est.get('telefono', 'No especificado')
            })
        
        # Usar la nueva función SIN pandas
        excel_bytes = descargar_excel_como_bytes(data, "reporte_estudiantes.xlsx")
        
        # Enviar archivo
        return Response(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=reporte_estudiantes.xlsx",
                "Content-type": "application/vnd.ms-excel"
            }
        )
        
    except Exception as e:
        print(f"❌ Error generando Excel: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/descargar/pdf')
def descargar_pdf():
    """Descargar reporte de estudiantes en PDF SIN pandas"""
    try:
        # Obtener todos los estudiantes
        estudiantes = db.get_todos_estudiantes(limit=1000)
        
        if not estudiantes:
            return jsonify({'success': False, 'error': 'No hay estudiantes para exportar'})
        
        # Preparar datos para PDF
        data = []
        for est in estudiantes:
            data.append({
                'Matrícula': est['matricula'],
                'Nombre': f"{est['nombre']} {est['apellido']}",
                'Carrera': est['carrera'],
                'Semestre': est['semestre'],
                'Fecha': est['fecha_inscripcion'].strftime('%Y-%m-%d') if est['fecha_inscripcion'] else 'N/A',
                'Estado': 'PAGADO' if est.get('inscripcion_pagada') else 'PENDIENTE'
            })
        
        # Usar la nueva función SIN pandas
        pdf_buffer = generar_pdf_sin_pandas(data, "REPORTE DE ESTUDIANTES - UNIVERSIDAD")
        
        # Enviar archivo
        return Response(
            pdf_buffer.getvalue(),
            mimetype="application/pdf",
            headers={
                "Content-Disposition": "attachment; filename=reporte_estudiantes.pdf"
            }
        )
        
    except Exception as e:
        print(f"❌ Error generando PDF: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/descargar/reporte/pendientes')
def descargar_reporte_pendientes():
    """Descargar reporte específico de estudiantes pendientes SIN pandas"""
    try:
        estudiantes = db.get_estudiantes_pendientes_inscripcion()
        
        if not estudiantes:
            return jsonify({'success': False, 'error': 'No hay estudiantes pendientes'})
        
        # Preparar datos
        data = []
        for est in estudiantes:
            data.append({
                'Matrícula': est['matricula'],
                'Nombre': f"{est['nombre']} {est['apellido']}",
                'Carrera': est['carrera'],
                'Semestre': est['semestre'],
                'Fecha Inscripción': est['fecha_inscripcion'].strftime('%Y-%m-%d') if est['fecha_inscripcion'] else 'N/A',
                'Email': est.get('email', 'No especificado'),
                'Teléfono': est.get('telefono', 'No especificado')
            })
        
        # Usar la nueva función SIN pandas
        excel_bytes = descargar_excel_como_bytes(data, "estudiantes_pendientes.xlsx")
        
        return Response(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=estudiantes_pendientes.xlsx"
            }
        )
        
    except Exception as e:
        print(f"❌ Error generando reporte pendientes: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/diagnostico')
def diagnostico():
    """Endpoint temporal para diagnóstico"""
    try:
        resultado = db.diagnosticar_estudiantes()
        return jsonify({
            'success': True,
            'diagnostico': resultado
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    # Render asigna puerto via variable de entorno
    port = int(os.environ.get('PORT', 5000))
    
    # En producción usar gunicorn, en desarrollo el servidor de Flask
    if os.environ.get('FLASK_ENV') == 'production':
        app.run(host='0.0.0.0', port=port)
    else:
        app.run(host='0.0.0.0', port=port, debug=True)